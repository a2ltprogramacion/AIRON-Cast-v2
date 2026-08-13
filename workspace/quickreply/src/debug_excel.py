import sys, os
sys.path.insert(0, 'Y:/Proyectos/AIRON-Cast/workspace/quickreply/src')
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'quickreply.settings')

import pandas as pd
from openpyxl import load_workbook

for fname in ['quickreply_precios (formulador).xlsx', 'quickreply_precios.xlsx']:
    path = 'Y:/Proyectos/AIRON-Cast/workspace/quickreply/docs/temp/' + fname
    print('=== ' + fname + ' ===')
    try:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
        print('Hojas: ' + str(wb.sheetnames))
        wb.close()
    except Exception as e:
        print('Error abrir: ' + str(e))
        continue

    try:
        df = pd.read_excel(path, sheet_name=0, dtype=str, header=0)
        print('Filas: ' + str(len(df)))
        print('Columnas: ' + str(list(df.columns[:8])))
        first_row = df.iloc[0].tolist() if len(df) > 0 else []
        print('Primera fila: ' + str(first_row[:6]))
        print()
    except Exception as e:
        print('Error leer: ' + str(e))
        print()