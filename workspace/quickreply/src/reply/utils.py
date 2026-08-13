"""
Utils para QuickReply: carga de Excel y motor de renderizado de templates.
"""
import re
import pandas as pd
from openpyxl import load_workbook


def normalizar_columna(nombre):
    """Limpia el nombre de una columna: elimina saltos de linea, espacios extra y pasa a mayusculas."""
    if not isinstance(nombre, str):
        return ""
    return re.sub(r"[\n\r\t]+", " ", nombre).strip().upper()


def cargar_excel(ruta_archivo):
    """
    Carga productos desde un archivo Excel.

    Nombre de hoja por defecto: G3 Multi (configurable via settings si se necesita).
    Columnas esperadas (sin saltos de linea, normalizadas a mayusculas):
        - TIPO
        - CODIGO
        - PRODUCTO
        - PRECIO_USD: cualquier columna que contenga "PRECIO" y "PUBLICADO" (excepto BCV)
        - PRECIO_BCV: cualquier columna que contenga "PRECIO" y "BCV"

    Usa data_only=True para forzar lectura de valores calculados (no formulas).
    Si la hoja no existe, intenta con 'G3 Multi - Ajuste' o muestra todas las hojas disponibles.

    Retorna:
        dict con claves: 'creados', 'actualizados', 'errores', 'total', 'debug'
    """
    from reply.models import Product

    stats = {
        "creados": 0,
        "actualizados": 0,
        "errores": [],
        "total": 0,
        "debug": {},
    }

    try:
        wb = load_workbook(filename=ruta_archivo, data_only=True, read_only=True)
        hojas_disponibles = wb.sheetnames
        wb.close()
    except Exception as e:
        stats["errores"].append(f"No se pudo abrir el archivo Excel: {e}")
        return stats

    stats["debug"]["hojas_encontradas"] = hojas_disponibles

    hoja_nombre = _detectar_hoja_precios(hojas_disponibles)
    if hoja_nombre is None:
        stats["errores"].append(
            f"No se encontro hoja de precios. Hojas disponibles: {hojas_disponibles}"
        )
        return stats

    stats["debug"]["hoja_usada"] = hoja_nombre

    try:
        df = pd.read_excel(
            ruta_archivo,
            sheet_name=hoja_nombre,
            dtype=str,
            engine="openpyxl",
        )
    except Exception as e:
        stats["errores"].append(f"Error al leer la hoja '{hoja_nombre}': {e}")
        return stats

    if df.empty:
        stats["errores"].append(
            f"La hoja '{hoja_nombre}' esta vacia. "
            f"Verifica que el archivo haya sido guardado con los calculos actualizados "
            f"(guarda con Ctrl+Shift+F9 en Excel para forzar recALCULO)."
        )
        return stats

    stats["debug"]["filas_en_hoja"] = len(df)
    df.columns = [normalizar_columna(c) for c in df.columns]
    stats["debug"]["columnas_detectadas"] = list(df.columns)

    COL_CODIGO = "CODIGO"
    COL_PRODUCTO = "PRODUCTO"
    COL_TIPO = "TIPO"

    col_usd = _detectar_columna_precio(df.columns, "usd")
    col_bcv = _detectar_columna_precio(df.columns, "bcv")

    if col_usd is None:
        stats["errores"].append(
            f"No se encontro columna de precio USD. "
            f"Columnas disponibles: {list(df.columns)}"
        )
        return stats
    if col_bcv is None:
        stats["errores"].append(
            f"No se encontro columna de precio BCV. "
            f"Columnas disponibles: {list(df.columns)}"
        )
        return stats

    stats["debug"]["col_usd"] = col_usd
    stats["debug"]["col_bcv"] = col_bcv

    for idx, row in df.iterrows():
        try:
            codigo = str(row.get(COL_CODIGO, "")).strip()
            if not codigo or codigo.lower() in ("nan", "none", ""):
                continue

            producto = str(row.get(COL_PRODUCTO, "")).strip()
            if producto.lower() in ("nan", "none"):
                producto = ""

            tipo = str(row.get(COL_TIPO, "")).strip()
            if tipo.lower() in ("nan", "none", ""):
                tipo = None

            precio_usd_raw = row.get(col_usd, None)
            precio_bcv_raw = row.get(col_bcv, None)

            try:
                precio_usd = float(pd.to_numeric(precio_usd_raw, errors="coerce") or 0)
            except (ValueError, TypeError):
                precio_usd = 0.0

            try:
                precio_bcv = float(pd.to_numeric(precio_bcv_raw, errors="coerce") or 0)
            except (ValueError, TypeError):
                precio_bcv = 0.0

            if precio_usd == 0 and precio_bcv == 0:
                stats["errores"].append(
                    f"Fila {idx + 2}: Codigo '{codigo}' tiene precios en cero (posible formula sin calcular)"
                )
                continue

            obj, created = Product.objects.update_or_create(
                codigo=codigo,
                defaults={
                    "producto": producto,
                    "tipo": tipo,
                    "precio_usd": precio_usd,
                    "precio_bcv": precio_bcv,
                },
            )

            if created:
                stats["creados"] += 1
            else:
                stats["actualizados"] += 1

            stats["total"] += 1

        except Exception as e:
            stats["errores"].append(f"Fila {idx + 2}: {e}")

    return stats


def _detectar_hoja_precios(hojas):
    """
    Detecta la hoja de precios entre una lista de nombres de hojas.
    Preferencias: 'G3 Multi', luego cualquier otra que parezca de precios.
    Compara stripping whitespace para evitar problemas con nombres como 'G3 Multi '.
    """
    prefs = ["G3 Multi", "G3 Multi - Ajuste", "Precios", "precios", "LISTA DE PRECIOS"]
    hojas_norm = [h.strip() for h in hojas]
    prefs_norm = [p.strip() for p in prefs]

    for pref_norm, pref_orig in zip(prefs_norm, prefs):
        for hoja_norm, hoja_orig in zip(hojas_norm, hojas):
            if hoja_norm == pref_norm:
                return hoja_orig

    for hoja_norm, hoja_orig in zip(hojas_norm, hojas):
        h = hoja_norm.upper()
        if any(k in h for k in ["G3", "PRECIO", "PRODUCTO", "ARTICULO", "INVENTARIO"]):
            return hoja_orig

    return hojas[0] if hojas else None


def _detectar_columna_precio(columnas, tipo):
    """
    Detecta la columna de precio adecuada.
    USD: contiene 'PRECIO' y 'PUBLICADO' pero NO 'BCV'
    BCV: contiene 'PRECIO' y 'BCV'
    """
    for col in columnas:
        if tipo == "usd":
            if "PRECIO" in col and "PUBLICADO" in col and "BCV" not in col:
                return col
        elif tipo == "bcv":
            if "PRECIO" in col and "BCV" in col:
                return col
    return None


def render_template(contenido):
    """
    Renderiza un mensaje de plantilla sustituyendo tokens {CODIGO_usd} y {CODIGO_bcv}
    por los precios correspondientes de la base de datos.

    Formato de tokens:
        {CODIGO_usd}  -> precio en dolares (ej: 45.0)
        {CODIGO_bcv}  -> precio paralelo formateado (ej: 70,00)

    Si el CODIGO no existe en la base de datos, el token se reemplaza por:
        [AGOTADO / CONSULTAR]
    """
    from reply.models import Product

    def reemplazar_token_usd(match):
        codigo = match.group(1).strip().upper()
        try:
            producto = Product.objects.get(codigo=codigo)
            return f"{producto.precio_usd:.2f}"
        except Product.DoesNotExist:
            return "[AGOTADO / CONSULTAR]"

    def reemplazar_token_bcv(match):
        codigo = match.group(1).strip().upper()
        try:
            producto = Product.objects.get(codigo=codigo)
            return f"{producto.precio_bcv:,.2f}"
        except Product.DoesNotExist:
            return "[AGOTADO / CONSULTAR]"

    resultado = re.sub(
        r"\{(\w+)_usd\}", reemplazar_token_usd, contenido, flags=re.IGNORECASE
    )
    resultado = re.sub(
        r"\{(\w+)_bcv\}", reemplazar_token_bcv, resultado, flags=re.IGNORECASE
    )

    return resultado


def formatear_bcv(valor):
    """Formatea un valor numerico como precio BCV con separador de miles y 2 decimales."""
    try:
        return f"{float(valor):,.2f}"
    except (ValueError, TypeError):
        return "0,00"